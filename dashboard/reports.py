"""Excel trade-report generation, queried straight from the `trades` table."""

import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl.utils import get_column_letter

REPORT_COLUMNS: List[str] = [
    "Date",
    "Symbol",
    "Market",
    "Strategy",
    "Side",
    "Entry Price",
    "Exit Price",
    "Quantity",
    "P&L",
    "P&L %",
    "Duration",
    "Exit Reason",
    "Parameters",
]


def generate_excel_report(
    db_path: str = "data/trading_agent.db",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    output_path: Optional[str] = None,
    reports_dir: str = "data/reports",
) -> str:
    """Build an Excel report of closed trades in [start_date, end_date] and
    return the path written to."""
    trades_df = _load_trades(db_path, start_date, end_date)
    report_df = _format_report(trades_df)

    if output_path is None:
        output_path = _default_output_path(reports_dir, start_date, end_date)
    else:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    report_df.to_excel(output_path, index=False, sheet_name="Trades")
    _autofit_columns(output_path)

    return output_path


def _load_trades(db_path: str, start_date: Optional[str], end_date: Optional[str]) -> pd.DataFrame:
    conditions = ["status = 'CLOSED'"]
    params: List[str] = []
    if start_date:
        conditions.append("exit_time >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("exit_time <= ?")
        params.append(end_date)

    query = f"SELECT * FROM trades WHERE {' AND '.join(conditions)} ORDER BY exit_time"
    conn = sqlite3.connect(db_path)
    try:
        return pd.read_sql_query(query, conn, params=params)
    finally:
        conn.close()


def _format_report(trades_df: pd.DataFrame) -> pd.DataFrame:
    if trades_df.empty:
        return pd.DataFrame(columns=REPORT_COLUMNS)

    return pd.DataFrame(
        {
            "Date": pd.to_datetime(trades_df["exit_time"]).dt.date.astype(str),
            "Symbol": trades_df["symbol"],
            "Market": trades_df["market"],
            "Strategy": trades_df["strategy_name"],
            "Side": trades_df["side"],
            "Entry Price": trades_df["entry_price"],
            "Exit Price": trades_df["exit_price"],
            "Quantity": trades_df["quantity"],
            "P&L": trades_df["pnl"],
            "P&L %": trades_df["pnl_pct"],
            "Duration": trades_df["duration_minutes"],
            "Exit Reason": trades_df["exit_reason"],
            "Parameters": trades_df["parameters"],
        }
    )


def _default_output_path(reports_dir: str, start_date: Optional[str], end_date: Optional[str]) -> str:
    Path(reports_dir).mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    range_part = f"{_sanitize(start_date) or 'all'}_{_sanitize(end_date) or 'all'}"
    return str(Path(reports_dir) / f"trades_report_{range_part}_{timestamp}.xlsx")


def _sanitize(value: Optional[str]) -> Optional[str]:
    if not value:
        return value
    return re.sub(r"[^0-9A-Za-z_-]", "-", value)


def _autofit_columns(path: str) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook["Trades"]
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        sheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    workbook.save(path)
